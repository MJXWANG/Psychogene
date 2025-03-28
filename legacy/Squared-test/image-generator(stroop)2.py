import openpyxl
from tkinter import Tk
from tkinter.filedialog import askopenfilename, askdirectory
import matplotlib.pyplot as plt

# 定义背景颜色（归一化RGB值）
bg_color = (200/255, 200/255, 200/255)

# 隐藏 Tkinter 主窗口
root = Tk()
root.withdraw()

# 选择 Excel 文件
excel_path = askopenfilename(title="请选择 Excel 文件", filetypes=[("Excel files", "*.xlsx;*.xls")])
if not excel_path:
    print("未选择 Excel 文件，程序退出。")
    exit()

# 选择图片保存的文件夹
output_folder = askdirectory(title="请选择保存图片的文件夹")
if not output_folder:
    print("未选择保存路径，程序退出。")
    exit()

# 询问行号范围（假设第一行为表头，数据从第二行开始）
start_row = int(input("请输入起始行号（从2开始，假设第1行为表头）："))
end_row = int(input("请输入结束行号："))

# 询问生成图片所用的三个列名称（用逗号分隔，例如：Col1, Col2, Col3）
columns_input = input("请输入生成图片的三个列名称（用逗号分隔）：")
cols_headers = [col.strip() for col in columns_input.split(",")]
if len(cols_headers) != 3:
    print("错误：请输入恰好三个列名称。")
    exit()

# 加载工作簿和活动工作表
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# 建立表头名称到列字母的映射（假设表头在第1行）
header_to_col = {}
for cell in ws[1]:
    if cell.value is not None:
        header_to_col[str(cell.value).strip()] = cell.column_letter

# 检查输入的列名称是否在表头中
for header in cols_headers:
    if header not in header_to_col:
        print(f"错误：在Excel表头中未找到 '{header}'")
        exit()

# 定义一个函数，读取单元格字体颜色，返回标准的十六进制颜色代码（如 "#RRGGBB"）
def get_font_color(cell):
    color_hex = "#000000"  # 默认黑色
    if cell.font and cell.font.color and cell.font.color.rgb:
        try:
            rgb_val = cell.font.color.rgb
            # 如果不是字符串，则转换为字符串
            if not isinstance(rgb_val, str):
                rgb_str = str(rgb_val)
            else:
                rgb_str = rgb_val
            # 若为8位（AARRGGBB格式），去掉前两位Alpha值
            if len(rgb_str) == 8 and all(ch in "0123456789ABCDEFabcdef" for ch in rgb_str):
                color_hex = "#" + rgb_str[2:]
            # 若为6位直接使用
            elif len(rgb_str) == 6 and all(ch in "0123456789ABCDEFabcdef" for ch in rgb_str):
                color_hex = "#" + rgb_str
        except Exception as e:
            color_hex = "#000000"
    return color_hex

# 初始化计数器
counter = 1

# 针对指定行范围，每一行生成一张图片
for r in range(start_row, end_row + 1):
    # 根据表头映射，获取对应的单元格（构造单元格引用，如 "B2"）
    cell_top = ws[f"{header_to_col[cols_headers[0]]}{r}"]
    cell_bottom_left = ws[f"{header_to_col[cols_headers[1]]}{r}"]
    cell_bottom_right = ws[f"{header_to_col[cols_headers[2]]}{r}"]

    text_top = str(cell_top.value) if cell_top.value is not None else ""
    text_bottom_left = str(cell_bottom_left.value) if cell_bottom_left.value is not None else ""
    text_bottom_right = str(cell_bottom_right.value) if cell_bottom_right.value is not None else ""

    color_top = get_font_color(cell_top)
    color_bottom_left = get_font_color(cell_bottom_left)
    color_bottom_right = get_font_color(cell_bottom_right)

    # 创建图像，设置背景为bg_color，尺寸500x300像素（figsize=(5,3) 配合 dpi=100）
    plt.figure(figsize=(5, 3), facecolor=bg_color, dpi=100)
    plt.axis('off')

    # 放置文本：中上、右边、左边（根据需求调整位置）
    plt.text(0.5, 0.70, text_top, ha='center', va='center', fontsize=30, color=color_top)
    plt.text(0.75, 0.40, text_bottom_left, ha='center', va='center', fontsize=25, color=color_bottom_left)
    plt.text(0.25, 0.40, text_bottom_right, ha='center', va='center', fontsize=25, color=color_bottom_right)

    # 构造输出文件名，使用顺序号和三个文本内容（并替换非法字符）
    filename = f"_{counter}_{text_top.strip()}_{text_bottom_left.strip()}_{text_bottom_right.strip()}.png"
    for ch in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
        filename = filename.replace(ch, '_')
    save_path = output_folder + "/" + filename

    # 保存图片，确保背景颜色为bg_color
    plt.savefig(save_path, bbox_inches='tight', pad_inches=0.1, facecolor=bg_color)
    plt.close()

    # 打印文件名（去掉 .png 后缀），每个名称单独一行
    base_filename = filename.replace(".png", "")
    print(base_filename)

    # 更新计数器
    counter += 1
